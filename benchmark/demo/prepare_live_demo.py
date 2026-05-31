#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Prepare synthetic files for a real Phase F screen-recording demo."""

from __future__ import annotations

import argparse
import csv
import json
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUT_DIR = REPO_ROOT / "_system" / "data" / "phase_f_live_demo"
SHEET_NAME = "最終ナレッジ候補一覧"
HEADERS = [
    "ナレッジID",
    "グループID",
    "候補_質問",
    "候補_回答",
    "カテゴリ",
    "統合件数",
    "既存FAQ_ID",
    "既存FAQ_質問",
    "既存FAQ_回答",
    "既存FAQ_比較",
    "リスクレベル",
    "信頼度",
    "推奨アクション",
    "判定根拠",
    "レビュー結果",
]


def _write_inquiry_log(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = [
        {
            "ticket_id": "SYN-001",
            "question": "VPNにつながりません",
            "answer": "最新版VPNクライアントへ更新してから接続してください。",
            "category": "ITサポート",
            "note": "Synthetic inquiry log. Not customer data.",
        },
        {
            "ticket_id": "SYN-002",
            "question": "承認済みになっていない申請期限を教えて",
            "answer": "未承認のため担当部門確認が必要です。",
            "category": "申請",
            "note": "Synthetic no-match inquiry. Not customer data.",
        },
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="EAF1FF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {
        "A": 16,
        "B": 14,
        "C": 28,
        "D": 54,
        "E": 16,
        "F": 10,
        "G": 18,
        "H": 28,
        "I": 44,
        "J": 24,
        "K": 14,
        "L": 10,
        "M": 22,
        "N": 58,
        "O": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def _write_review_excel(path: Path, review_result: str = "未確認") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    ws.append(
        [
            "draft-vpn-update-001",
            "grp-vpn-001",
            "VPNにつながりません",
            "最新版VPNクライアントへ更新してから接続してください。更新後も接続できない場合は、端末を再起動してから再度接続を試してください。",
            "ITサポート",
            "3",
            "demo-vpn-001",
            "VPNにつながりません",
            "VPNクライアントを再起動してから、もう一度接続してください。",
            "既存FAQと内容が近いが、回答手順に更新差分あり",
            "low",
            "0.92",
            "既存FAQ更新",
            "P3-2確認（既存FAQ更新候補）。既存FAQの質問に一致し、回答内容が最新版VPNクライアントへの更新手順へ変わっているため、人間レビューで更新可否を確認する。",
            review_result,
        ]
    )
    _style_sheet(ws)
    dv = DataValidation(
        type="list",
        formula1='"未確認,採用,差し戻し"',
        allow_blank=False,
        showErrorMessage=True,
    )
    ws.add_data_validation(dv)
    dv.add("O2:O20")

    raw = wb.create_sheet("元ログ")
    raw.append(["ticket_id", "question", "answer", "category", "note"])
    raw.append(
        [
            "SYN-001",
            "VPNにつながりません",
            "最新版VPNクライアントへ更新してから接続してください。",
            "ITサポート",
            "Synthetic inquiry log. Not customer data.",
        ]
    )
    raw.append(
        [
            "SYN-002",
            "承認済みになっていない申請期限を教えて",
            "未承認のため担当部門確認が必要です。",
            "申請",
            "Synthetic no-match inquiry. Not customer data.",
        ]
    )
    wb.save(path)


def prepare(out_dir: Path) -> dict[str, str]:
    out_dir.mkdir(parents=True, exist_ok=True)
    inquiry_log = out_dir / "01_synthetic_inquiry_log.csv"
    review_excel = out_dir / "02_FAQ_final_result_for_review.xlsx"
    approved_excel = out_dir / "03_FAQ_final_result_approved.xlsx"
    before_json = out_dir / "approved_knowledge_before.json"
    merged_json = out_dir / "approved_knowledge_merged.json"

    _write_inquiry_log(inquiry_log)
    _write_review_excel(review_excel, review_result="未確認")
    _write_review_excel(approved_excel, review_result="採用")
    shutil.copyfile(
        REPO_ROOT / "benchmark" / "demo" / "approved_knowledge_before.json",
        before_json,
    )
    if merged_json.exists():
        merged_json.unlink()

    manifest = {
        "synthetic_data": "true",
        "inquiry_log": str(inquiry_log),
        "review_excel": str(review_excel),
        "approved_excel": str(approved_excel),
        "before_json": str(before_json),
        "merged_json": str(merged_json),
    }
    with (out_dir / "manifest.json").open("w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare Phase F live demo fixtures.")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()
    manifest = prepare(args.out_dir)
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
