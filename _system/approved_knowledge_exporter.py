#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
レビュー済みExcelから承認済みKnowledgeを出力する。
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


SHEET_NAME = "最終ナレッジ候補一覧"
APPROVED_REVIEW_RESULT = "採用"
JST = timezone(timedelta(hours=9))


def _cell_text(value: Any) -> str:
    """Excelセル値をJSON向け文字列に整える。"""
    if value is None:
        return ""
    return str(value).strip()


def _build_header_index(ws) -> Dict[str, int]:
    """1行目のヘッダー名から列番号へのマップを作る。"""
    headers: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        header = _cell_text(cell.value)
        if header:
            headers[header] = col_idx
    return headers


def _required_value(ws, row_idx: int, headers: Dict[str, int], header: str) -> str:
    """必須列の値を取得する。"""
    if header not in headers:
        raise ValueError(f"必須列が見つかりません: {header}")
    return _cell_text(ws.cell(row=row_idx, column=headers[header]).value)


def load_approved_knowledge_from_excel(
    excel_path_or_file: str | Path | Any,
    approved_at: str | None = None,
) -> List[Dict[str, str]]:
    """レビュー結果が「採用」の行だけ approved_knowledge 形式で返す。"""
    wb = load_workbook(excel_path_or_file, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {SHEET_NAME}")

    ws = wb[SHEET_NAME]
    headers = _build_header_index(ws)
    approved_at_value = approved_at or datetime.now(JST).isoformat(timespec="seconds")

    approved_items: List[Dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        review_result = _required_value(ws, row_idx, headers, "レビュー結果")
        if review_result != APPROVED_REVIEW_RESULT:
            continue

        approved_items.append(
            {
                "knowledge_id": _required_value(ws, row_idx, headers, "ナレッジID"),
                "question": _required_value(ws, row_idx, headers, "候補_質問"),
                "answer": _required_value(ws, row_idx, headers, "候補_回答"),
                "category": _required_value(ws, row_idx, headers, "カテゴリ"),
                "approved_status": "approved",
                "approved_at": approved_at_value,
            }
        )

    return approved_items


def export_approved_knowledge_from_excel(
    excel_path_or_file: str | Path | Any,
    output_path: str | Path,
    approved_at: str | None = None,
) -> Path:
    """レビュー済みExcelから approved_knowledge.json を出力する。"""
    approved_items = load_approved_knowledge_from_excel(
        excel_path_or_file=excel_path_or_file,
        approved_at=approved_at,
    )
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8") as f:
        json.dump(approved_items, f, ensure_ascii=False, indent=2)
    return target


def main() -> None:
    parser = argparse.ArgumentParser(
        description="レビュー済みExcelから approved_knowledge.json を出力します。"
    )
    parser.add_argument("excel_path", help="FAQ_final_result.xlsx のパス")
    parser.add_argument(
        "-o",
        "--output",
        default="data/outputs/approved_knowledge.json",
        help="出力先JSONパス",
    )
    args = parser.parse_args()

    output_path = export_approved_knowledge_from_excel(
        excel_path_or_file=args.excel_path,
        output_path=args.output,
    )
    print(f"approved_knowledge出力: {output_path}")


if __name__ == "__main__":
    main()
