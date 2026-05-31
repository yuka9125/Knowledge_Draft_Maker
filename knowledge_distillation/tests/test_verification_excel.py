#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""最終レビューExcel出力の単体テスト。"""

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from knowledge_distillation.verification_excel import VerificationExcelWriter


class VerificationExcelWriterTest(unittest.TestCase):
    """Sheet1の候補参考行を検証する。"""

    def test_answer_candidate_rows_are_written_as_reference_rows(self):
        tmp_dir = Path(tempfile.mkdtemp(prefix="verification_excel_test_"))
        writer = VerificationExcelWriter(
            output_dir=str(tmp_dir),
            phase="FAQ_final_result",
            timestamp="20260531_000000",
        )
        writer.add_final_faq_sheet(
            sheet_name="最終ナレッジ候補一覧",
            groups=[],
            knowledge_candidates=[
                {
                    "knowledge_id": "k-001",
                    "group_id": "g-001",
                    "question": "VPNに接続できない場合は？",
                    "answer": "最新版へ更新してください。",
                    "category": "IT",
                    "similar_logs_count": 3,
                    "risk_level": "low",
                    "confidence": 0.95,
                    "recommended_action": "新規FAQ作成",
                    "judgement_reason": "判定根拠",
                    "review_result": "",
                    "answer_candidate_2": "再起動してください。",
                    "answer_candidate_3": "ネットワークを切り替えてください。",
                }
            ],
        )
        path = writer.save()

        wb = load_workbook(path, data_only=True)
        ws = wb["最終ナレッジ候補一覧"]
        headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

        self.assertEqual(ws.max_row, 4)
        self.assertEqual(ws.cell(2, headers["ナレッジID"]).value, "k-001")
        self.assertEqual(ws.cell(2, headers["レビュー結果"]).value, "未確認")
        self.assertEqual(ws.cell(3, headers["ナレッジID"]).value, None)
        self.assertEqual(
            ws.cell(3, headers["候補_回答"]).value,
            "再起動してください。",
        )
        self.assertEqual(ws.cell(3, headers["レビュー結果"]).value, "候補参考")
        self.assertEqual(
            ws.cell(4, headers["候補_回答"]).value,
            "ネットワークを切り替えてください。",
        )
        self.assertEqual(ws.cell(4, headers["レビュー結果"]).value, "候補参考")

        validations = list(ws.data_validations.dataValidation)
        self.assertTrue(validations)
        self.assertIn("候補参考", validations[0].formula1)


if __name__ == "__main__":
    unittest.main()
