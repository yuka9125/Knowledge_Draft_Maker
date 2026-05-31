#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
approved_knowledge出力の単体テスト。
"""

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from knowledge_distillation.approved_knowledge_exporter import (
    export_approved_knowledge_from_excel,
    load_approved_knowledge_from_excel,
    merge_approved_knowledge,
)


HEADERS = [
    "ナレッジID",
    "グループID",
    "候補_質問",
    "候補_回答",
    "カテゴリ",
    "統合件数",
    "既存FAQ_ID",
    "既存FAQ_質問",
    "既存FAQ_回答",
    "既存FAQ_比較",
    "リスクレベル",
    "信頼度",
    "推奨アクション",
    "判定根拠",
    "レビュー結果",
]


def _row(
    knowledge_id="",
    question="",
    answer="",
    category="IT",
    existing_faq_id="",
    action="新規FAQ作成",
    review="採用",
):
    return [
        knowledge_id, 1, question, answer, category, 1,
        existing_faq_id, "", "", "既存FAQなし/未照合",
        "low", 0.9, action, "", review,
    ]


class ApprovedKnowledgeExporterTest(unittest.TestCase):
    """レビュー済みExcelから採用行だけ出力することを検証。"""

    def _create_reviewed_excel(self) -> Path:
        tmp_dir = Path(tempfile.mkdtemp(prefix="approved_knowledge_test_"))
        xlsx_path = tmp_dir / "FAQ_final_result.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "最終ナレッジ候補一覧"
        ws.append(
            [
                "ナレッジID",
                "グループID",
                "候補_質問",
                "候補_回答",
                "カテゴリ",
                "統合件数",
                "既存FAQ_ID",
                "既存FAQ_質問",
                "既存FAQ_回答",
                "既存FAQ_比較",
                "リスクレベル",
                "信頼度",
                "推奨アクション",
                "判定根拠",
                "レビュー結果",
            ]
        )
        ws.append(
            [
                "k-001",
                1,
                "VPNにつながりません",
                "最新版VPNクライアントへ更新してください。",
                "IT",
                2,
                "",
                "",
                "",
                "既存FAQなし/未照合",
                "low",
                0.91,
                "新規FAQ作成",
                "信頼度理由: 0.91\nリスク理由: low",
                "採用",
            ]
        )
        ws.append(
            [
                "k-002",
                2,
                "未確認の質問",
                "未確認の回答",
                "IT",
                1,
                "",
                "",
                "",
                "既存FAQなし/未照合",
                "low",
                0.70,
                "新規FAQ作成",
                "",
                "未確認",
            ]
        )
        ws.append(
            [
                "k-003",
                3,
                "不採用の質問",
                "不採用の回答",
                "IT",
                1,
                "",
                "",
                "",
                "既存FAQなし/未照合",
                "low",
                0.70,
                "新規FAQ作成",
                "",
                "不採用",
            ]
        )
        ws.append(
            [
                "",
                "",
                "",
                "候補参考の回答",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "候補参考",
            ]
        )
        wb.save(xlsx_path)
        return xlsx_path

    def test_load_only_approved_rows(self):
        """レビュー結果が採用の行だけ返す。"""
        xlsx_path = self._create_reviewed_excel()
        items = load_approved_knowledge_from_excel(
            xlsx_path,
            approved_at="2026-05-29T10:00:00+09:00",
        )

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["knowledge_id"], "k-001")
        self.assertEqual(items[0]["question"], "VPNにつながりません")
        self.assertEqual(items[0]["answer"], "最新版VPNクライアントへ更新してください。")
        self.assertEqual(items[0]["category"], "IT")
        self.assertEqual(items[0]["approved_status"], "approved")
        self.assertEqual(items[0]["approved_at"], "2026-05-29T10:00:00+09:00")

    def test_candidate_reference_rows_are_ignored(self):
        """候補参考行は approved_knowledge に出力しない。"""
        xlsx_path = self._create_reviewed_excel()
        items = load_approved_knowledge_from_excel(
            xlsx_path,
            approved_at="2026-05-29T10:00:00+09:00",
        )

        self.assertEqual(len(items), 1)
        self.assertNotIn("候補参考の回答", [item["answer"] for item in items])

    def test_blank_approved_rows_are_ignored(self):
        """空ID/空質問の行が誤って採用になってもスキップする。"""
        xlsx_path = self._create_reviewed_excel()
        wb = Workbook()
        ws = wb.active
        ws.title = "最終ナレッジ候補一覧"
        ws.append(HEADERS)
        ws.append(_row(knowledge_id="", question="", answer="候補だけ", review="採用"))
        ws.append(_row(knowledge_id="k-001", question="Q", answer="A", review="採用"))
        wb.save(xlsx_path)

        items = load_approved_knowledge_from_excel(
            xlsx_path,
            approved_at="2026-05-29T10:00:00+09:00",
        )

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["knowledge_id"], "k-001")

    def test_export_approved_knowledge_json(self):
        """approved_knowledge.json の形式を検証。"""
        xlsx_path = self._create_reviewed_excel()
        output_path = xlsx_path.parent / "approved_knowledge.json"

        export_approved_knowledge_from_excel(
            xlsx_path,
            output_path,
            approved_at="2026-05-29T10:00:00+09:00",
        )

        with output_path.open(encoding="utf-8") as f:
            data = json.load(f)

        self.assertEqual(
            list(data[0].keys()),
            [
                "knowledge_id",
                "question",
                "answer",
                "category",
                "approved_status",
                "approved_at",
            ],
        )
        self.assertEqual(len(data), 1)


class UpsertMergeTest(unittest.TestCase):
    """既存FAQ_ID をキーにした upsert マージを検証。"""

    def _excel(self, rows) -> Path:
        tmp_dir = Path(tempfile.mkdtemp(prefix="upsert_test_"))
        xlsx_path = tmp_dir / "FAQ_final_result.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "最終ナレッジ候補一覧"
        ws.append(HEADERS)
        for r in rows:
            ws.append(r)
        wb.save(xlsx_path)
        return xlsx_path

    def _base_file(self, dir_path: Path, items) -> Path:
        base = dir_path / "approved_knowledge.json"
        with base.open("w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False)
        return base

    def test_merge_function_upsert_and_append(self):
        base = [
            {"knowledge_id": "faq-100", "question": "旧Q", "answer": "旧A", "category": "IT"},
            {"knowledge_id": "faq-200", "question": "別Q", "answer": "別A", "category": "IT"},
        ]
        incoming = [
            {"knowledge_id": "faq-100", "question": "新Q", "answer": "新A", "category": "IT"},
            {"knowledge_id": "k-001", "question": "新規Q", "answer": "新規A", "category": "IT"},
        ]
        merged = merge_approved_knowledge(base, incoming)
        ids = [m["knowledge_id"] for m in merged]
        # 既存faq-100は上書き、faq-200は維持、k-001は末尾追加（重複なし）
        self.assertEqual(ids, ["faq-100", "faq-200", "k-001"])
        self.assertEqual(merged[0]["answer"], "新A")
        self.assertEqual(merged[1]["answer"], "別A")

    def test_update_candidate_overwrites_existing_id(self):
        # 既存FAQ_ID=faq-100 の更新候補は、自分のナレッジID(k-009)ではなくfaq-100に上書き
        xlsx = self._excel([
            _row(knowledge_id="k-009", question="VPN更新後の手順", answer="新手順",
                 existing_faq_id="faq-100", action="既存FAQ更新"),
        ])
        out = xlsx.parent / "approved_knowledge.json"
        self._base_file(xlsx.parent, [
            {"knowledge_id": "faq-100", "question": "VPN手順", "answer": "旧手順", "category": "IT"},
        ])
        export_approved_knowledge_from_excel(xlsx, out, approved_at="2026-05-30T00:00:00+09:00")
        data = json.loads(out.read_text(encoding="utf-8"))
        self.assertEqual(len(data), 1)  # 重複が増えない
        self.assertEqual(data[0]["knowledge_id"], "faq-100")
        self.assertEqual(data[0]["answer"], "新手順")

    def test_new_candidate_is_appended(self):
        xlsx = self._excel([
            _row(knowledge_id="k-001", question="新規Q", answer="新規A"),
        ])
        out = xlsx.parent / "approved_knowledge.json"
        self._base_file(xlsx.parent, [
            {"knowledge_id": "faq-100", "question": "既存Q", "answer": "既存A", "category": "IT"},
        ])
        export_approved_knowledge_from_excel(xlsx, out, approved_at="2026-05-30T00:00:00+09:00")
        data = json.loads(out.read_text(encoding="utf-8"))
        ids = [d["knowledge_id"] for d in data]
        self.assertEqual(ids, ["faq-100", "k-001"])

    def test_no_merge_keeps_snapshot(self):
        xlsx = self._excel([
            _row(knowledge_id="k-001", question="Q", answer="A"),
        ])
        out = xlsx.parent / "approved_knowledge.json"
        self._base_file(xlsx.parent, [
            {"knowledge_id": "faq-100", "question": "既存Q", "answer": "既存A", "category": "IT"},
        ])
        export_approved_knowledge_from_excel(
            xlsx, out, approved_at="2026-05-30T00:00:00+09:00", merge=False
        )
        data = json.loads(out.read_text(encoding="utf-8"))
        ids = [d["knowledge_id"] for d in data]
        self.assertEqual(ids, ["k-001"])  # baseは無視して上書き


if __name__ == "__main__":
    unittest.main()
