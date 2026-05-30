#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
検証Excel出力 共通モジュール

各Phaseで使用する検証Excel出力の共通機能を提供

★修正: Phase 1.5廃止（P1.5_類似度カラム削除）
★修正: 生_対応結果_文字数カラム追加
★修正: FAQ対象外ステータスの背景色:グレー
★修正: 最終FAQ一覧からFAQ除外データを除外
"""

import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional, Any
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from knowledge_distillation.display_labels import display_label


# =============================================================================
# スタイル定義
# =============================================================================
HEADER_FILL = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUB_HEADER_FILL = PatternFill(
    start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
)
ADOPTED_FILL = PatternFill(
    start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
)  # 緑（採用）
# 回答候補2/3の背景色（薄いオレンジ）
ANSWER_CANDIDATE_SUB_FILL = PatternFill(
    start_color="FBE5D6", end_color="FBE5D6", fill_type="solid"
)
DELETED_FILL = PatternFill(
    start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
)  # オレンジ（削除）
# FAQ対象外の背景色（グレー）
FAQ_EXCLUDED_FILL = PatternFill(
    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GROUP_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="medium"),
    bottom=Side(style="thin"),
)


# =============================================================================
# GIDトラッカー（Phase間でのGID統合を追跡）
# =============================================================================
class GIDTracker:
    """Phase間でのGID統合を追跡するクラス"""

    def __init__(self):
        # p0_gid → 最終的な統合先p0_gid
        self.gid_mapping: Dict[int, int] = {}

    def register(self, p0_gid: int):
        """新規GIDを登録（自分自身を指す）"""
        if p0_gid not in self.gid_mapping:
            self.gid_mapping[p0_gid] = p0_gid

    def merge(self, absorbed_gid: int, representative_gid: int):
        """absorbed_gid を representative_gid に統合"""
        if absorbed_gid == representative_gid:
            return

        # representative_gidの最終統合先を取得
        final_rep_gid = self.get_final_gid(representative_gid)

        # absorbed_gid に既に統合されていたGIDも含めて更新
        for gid in list(self.gid_mapping.keys()):
            if self.get_final_gid(gid) == absorbed_gid:
                self.gid_mapping[gid] = final_rep_gid

        self.gid_mapping[absorbed_gid] = final_rep_gid

    def get_final_gid(self, p0_gid: int) -> int:
        """最終的な統合先GIDを取得"""
        if p0_gid not in self.gid_mapping:
            return p0_gid

        # チェーンを辿って最終GIDを取得
        current = p0_gid
        visited = set()
        while current in self.gid_mapping and current not in visited:
            visited.add(current)
            next_gid = self.gid_mapping[current]
            if next_gid == current:
                break
            current = next_gid

        return current

    def update_all_records(self, records: Dict[int, "ProcessingRecord"]):
        """全レコードのfinal_gidを最終統合先に更新"""
        for record in records.values():
            if record.p0_gid is not None:
                record.final_gid = self.get_final_gid(record.p0_gid)


# =============================================================================
# データクラス
# =============================================================================
@dataclass
class ProcessingRecord:
    """処理履歴レコード"""

    original_idx: int
    p0_gid: Optional[int] = None  # Phase0のグループID（変更しない）
    p0_similarity: Optional[float] = (
        None  # Phase0の類似度（完全一致=1.00, 類似=0.90~, 短文="短文"相当はNone）
    )
    p2_similarity: Optional[float] = None  # Phase2の類似度
    p3_1_similarity: Optional[float] = None
    p3_2_similarity: Optional[float] = None
    final_gid: Optional[int] = None
    final_result: str = (
        ""  # "◯採用", "P0削除（完全一致）", "P0削除（類似）", "P0削除（短文）", "P2削除（完全一致）", "P2削除（類似）", "P3-1削除", "P3-2確認（...）", "FAQ対象外"
    )
    raw_overview: str = ""
    raw_response: str = ""
    question: str = ""
    answer: str = ""
    matched_faq_question: str = ""
    matched_faq_answer: str = ""
    matched_faq_id: str = ""
    matched_faq_row: Optional[int] = None
    confidence_score: float = 0.0


@dataclass
class GroupCandidate:
    """グループ内の候補データ"""

    original_idx: int
    rank: int  # 1, 2, 3... (回答文字数順)
    is_adopted: bool
    similarity: Optional[float]  # 代表との類似度（代表自身は None）
    raw_overview: str = ""
    raw_response: str = ""
    raw_response_length: int = 0
    question: str = ""
    answer: str = ""
    answer_length: int = 0
    category: str = ""
    keywords: str = ""
    link_names: str = ""
    user_role: str = ""
    confidence_score: float = 0.0


@dataclass
class DuplicateGroup:
    """重複グループ"""

    group_id: int
    candidates: List[GroupCandidate] = field(default_factory=list)

    def get_adopted_count(self) -> int:
        """採用された候補の数を返す"""
        return sum(1 for c in self.candidates if c.is_adopted)

    def get_representative(self) -> Optional[GroupCandidate]:
        """代表（ランク1）を返す"""
        for c in self.candidates:
            if c.rank == 1:
                return c
        return None


# =============================================================================
# スタイル適用関数
# =============================================================================
def apply_header_style(cell, fill=None):
    """ヘッダーセルにスタイルを適用"""
    cell.fill = fill if fill else HEADER_FILL
    cell.font = (
        HEADER_FONT
        if fill == HEADER_FILL or fill is None
        else Font(bold=True, size=10)
    )
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    cell.border = THIN_BORDER


def apply_cell_style(cell, wrap=True, fill=None, is_group_start=False):
    """データセルにスタイルを適用"""
    cell.alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=wrap
    )
    cell.border = GROUP_BORDER if is_group_start else THIN_BORDER
    if fill:
        cell.fill = fill


def set_column_widths(ws, widths: List[int]):
    """列幅を設定"""
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


# =============================================================================
# 検証Excel出力クラス
# =============================================================================
class VerificationExcelWriter:
    """検証Excel出力クラス"""

    def __init__(
        self, output_dir: str, phase: str, timestamp: Optional[str] = None
    ):
        """
        初期化

        Args:
            output_dir: 出力ディレクトリ
            phase: Phase名（"Phase0", "Phase2", "FAQ_final_result"）
            timestamp: タイムスタンプ（省略時は現在時刻）
        """
        self.output_dir = output_dir
        self.phase = phase
        self.timestamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
        self.wb = Workbook()
        # デフォルトシートを削除
        self.wb.remove(self.wb.active)

    def get_filename(self) -> str:
        """ファイル名を取得"""
        if self.phase == "FAQ_final_result":
            return f"FAQ_final_result_{self.timestamp}.xlsx"
        elif self.phase == "Phase0":
            return f"Phase0_verification_{self.timestamp}.xlsx"
        elif self.phase == "Phase2":
            return f"Phase2_verification_{self.timestamp}.xlsx"
        else:
            return f"{self.phase}_verification_{self.timestamp}.xlsx"

    def add_final_faq_sheet(
        self,
        sheet_name: str,
        groups: List[DuplicateGroup],
        knowledge_candidates: Optional[List[Dict[str, Any]]] = None,
        include_raw: bool = True,
        processing_records: Optional[Dict[int, ProcessingRecord]] = None,
    ):
        """
        最終FAQ一覧シートを追加
        FAQ除外データ（回答="-"）は含めない
        """
        ws = self.wb.create_sheet(sheet_name)

        # ナレッジ候補レイアウト（新スキーマ）
        if knowledge_candidates is not None:
            headers = [
                "ナレッジID",
                "グループID",
                "候補_質問",
                "候補_回答",
                "カテゴリ",
                "統合件数",
                "既存FAQ_ID",
                "既存FAQ_質問",
                "既存FAQ_回答",
                "既存FAQ_比較",
                "リスクレベル",
                "信頼度",
                "推奨アクション",
                "判定根拠",
                "レビュー結果",
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                apply_header_style(cell)

            row_idx = 2
            for item in knowledge_candidates:
                row_values = [
                    str(item.get("knowledge_id", "")),
                    str(item.get("group_id", "")),
                    str(item.get("question", "")),
                    str(item.get("answer", "")),
                    str(item.get("category", "")),
                    item.get("similar_logs_count", 0),
                    str(item.get("matched_faq_id", "")),
                    str(item.get("matched_faq_question", "")),
                    str(item.get("matched_faq_answer", "")),
                    str(item.get("existing_faq_comparison", "")),
                    str(item.get("risk_level", "")),
                    item.get("confidence", 0.0),
                    str(item.get("recommended_action", "")),
                    str(item.get("judgement_reason", "")),
                    str(item.get("review_result", "未確認") or "未確認"),
                ]

                row_fill = ADOPTED_FILL
                for col_idx, value in enumerate(row_values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    apply_cell_style(cell, fill=row_fill)
                row_idx += 1

            widths = [
                12,
                10,
                48,
                60,
                18,
                12,
                16,
                48,
                60,
                22,
                12,
                10,
                18,
                70,
                16,
            ]
            set_column_widths(ws, widths)

            # レビュー結果列は固定選択肢のみ入力可能
            if row_idx > 2:
                review_result_col = 15  # 1始まり
                dv = DataValidation(
                    type="list",
                    formula1='"未確認,採用,不採用"',
                    allow_blank=False,
                    showErrorMessage=True,
                    errorTitle="入力値エラー",
                    error="レビュー結果は指定された選択肢から入力してください。",
                )
                ws.add_data_validation(dv)
                dv.add(
                    f"{get_column_letter(review_result_col)}2:"
                    f"{get_column_letter(review_result_col)}{row_idx - 1}"
                )

            ws.row_dimensions[1].height = 25
            for row in range(2, row_idx):
                ws.row_dimensions[row].height = 60
            return

        # ヘッダー構成
        headers = [
            "No",
            "元idx",
            "グループID",
            "カテゴリ",
            "立場",
            "質問",
            "候補1_回答",
            "候補2_回答",
            "候補3_回答",
            "キーワード",
            "リンク名",
            "統合件数",
            "信頼度",
        ]
        if include_raw:
            headers.extend(["生_概要", "生_対応結果", "生_対応結果_文字数"])

        # 表示名を標準化
        headers = [display_label(h) for h in headers]

        # ヘッダー書き込み
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            apply_header_style(cell)

        # グループをgroup_idでソート
        sorted_groups = sorted(groups, key=lambda g: g.group_id)

        gid_counts = {}
        if processing_records:
            for record in processing_records.values():
                gid = record.final_gid
                if gid is not None:
                    gid_counts[gid] = gid_counts.get(gid, 0) + 1

        # データ書き込み
        row_idx = 2
        # データ書き込み
        row_idx = 2
        faq_no = 1
        for group in sorted_groups:
            for candidate in group.candidates:
                # 採用されたデータのみ、かつFAQ除外データを除外
                if candidate.is_adopted and candidate.answer != "-":
                    col_idx = 1

                    # No
                    cell = ws.cell(row=row_idx, column=col_idx, value=faq_no)
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # 元idxs
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=candidate.original_idx,
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # グループID
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=group.group_id
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # カテゴリ
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=candidate.category
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # 立場
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=candidate.user_role
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # 質問
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=candidate.question
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # 候補1_回答
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=candidate.answer
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # 候補2_回答、候補3_回答（不採用候補から取得、FAQ除外は除く）
                    other_answers = [
                        c.answer
                        for c in group.candidates
                        if not c.is_adopted and c.answer != "-"
                    ][:2]

                    # 候補2_回答
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=(
                            other_answers[0] if len(other_answers) > 0 else ""
                        ),
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # 候補3_回答
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=(
                            other_answers[1] if len(other_answers) > 1 else ""
                        ),
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # キーワード
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=candidate.keywords
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # リンク名
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=candidate.link_names
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # 統合件数
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=gid_counts.get(
                            group.group_id, len(group.candidates)
                        ),
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    # 信頼度
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=candidate.confidence_score,
                    )
                    apply_cell_style(cell, fill=ADOPTED_FILL)
                    col_idx += 1

                    if include_raw:
                        # 生_概要
                        cell = ws.cell(
                            row=row_idx,
                            column=col_idx,
                            value=candidate.raw_overview,
                        )
                        apply_cell_style(cell, fill=ADOPTED_FILL)
                        col_idx += 1

                        # 生_対応結果
                        cell = ws.cell(
                            row=row_idx,
                            column=col_idx,
                            value=candidate.raw_response,
                        )
                        apply_cell_style(cell, fill=ADOPTED_FILL)
                        col_idx += 1

                        # 生_対応結果_文字数
                        cell = ws.cell(
                            row=row_idx,
                            column=col_idx,
                            value=candidate.raw_response_length,
                        )
                        apply_cell_style(cell, fill=ADOPTED_FILL)
                        col_idx += 1

                    faq_no += 1
                    row_idx += 1

        # 列幅設定
        if include_raw:
            widths = [
                6,
                8,
                10,
                15,
                10,
                50,
                60,
                60,
                60,
                25,
                30,
                10,
                8,
                35,
                45,
                12,
            ]
        else:
            widths = [6, 8, 10, 15, 10, 50, 60, 60, 60, 25, 30, 10, 8]
        set_column_widths(ws, widths)

        # 行高さ設定
        ws.row_dimensions[1].height = 25
        for row in range(2, row_idx):
            ws.row_dimensions[row].height = 50

    def add_processing_history_sheet(
        self,
        sheet_name: str,
        records: List[ProcessingRecord],
    ):
        """
        全データ処理履歴シートを追加

        ★修正: P1.5_類似度カラムを削除
        ★修正: 生_対応結果_文字数カラムを追加
        ★修正: FAQ対象外の背景色を追加
        """
        ws = self.wb.create_sheet(sheet_name)

        # ヘッダー構成（P1.5_類似度削除、生_対応結果_文字数追加）
        headers = [
            "元idx",
            "P0_GID",
            "P0_類似度",
            "P2_類似度",
            "P3-1_類似度",
            "P3-2_類似度",
            "最終GID",
            "最終結果",
            "信頼度",
            "生_概要",
            "生_対応結果",
            "生_対応結果_文字数",  # ★追加
            "質問",
            "回答",
            "一致FAQ_行",
            "一致FAQ_質問",
            "一致FAQ_回答",
        ]

        # 表示名を標準化
        headers = [display_label(h) for h in headers]

        # ヘッダー書き込み
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            apply_header_style(cell)

        # レコードを最終GID、元idxでソート
        sorted_records = sorted(
            records, key=lambda r: (r.final_gid or 0, r.original_idx)
        )

        # データ書き込み
        current_gid = None
        for row_idx, record in enumerate(sorted_records, 2):
            # 採用/削除/FAQ対象外で背景色を変える
            if record.final_result == "◯採用":
                fill = ADOPTED_FILL
            elif record.final_result == "FAQ対象外":
                fill = FAQ_EXCLUDED_FILL
            elif "削除" in record.final_result:
                fill = DELETED_FILL
            else:
                fill = None

            # グループの先頭かどうか
            is_group_start = record.final_gid != current_gid
            current_gid = record.final_gid

            col_idx = 1

            # 元idx
            cell = ws.cell(
                row=row_idx, column=col_idx, value=record.original_idx
            )
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # P0_GID
            cell = ws.cell(row=row_idx, column=col_idx, value=record.p0_gid)
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # P0_類似度
            p0_val = (
                "−"
                if record.p0_similarity is None
                else f"{record.p0_similarity:.2f}"
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=p0_val)
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # P2_類似度
            p2_val = (
                "−"
                if record.p2_similarity is None
                else f"{record.p2_similarity:.2f}"
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=p2_val)
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # P3-1_類似度
            p3_1_val = (
                "−"
                if record.p3_1_similarity is None
                else f"{record.p3_1_similarity:.2f}"
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=p3_1_val)
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # P3-2_類似度
            p3_2_val = (
                "−"
                if record.p3_2_similarity is None
                else f"{record.p3_2_similarity:.2f}"
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=p3_2_val)
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 最終GID
            cell = ws.cell(row=row_idx, column=col_idx, value=record.final_gid)
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 最終結果
            cell = ws.cell(
                row=row_idx, column=col_idx, value=record.final_result
            )
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 信頼度
            cell = ws.cell(
                row=row_idx, column=col_idx, value=record.confidence_score
            )
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 生_概要
            cell = ws.cell(
                row=row_idx, column=col_idx, value=record.raw_overview
            )
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 生_対応結果
            cell = ws.cell(
                row=row_idx, column=col_idx, value=record.raw_response
            )
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 生_対応結果_文字数 ★追加
            raw_response_length = (
                len(record.raw_response) if record.raw_response else 0
            )
            cell = ws.cell(
                row=row_idx, column=col_idx, value=raw_response_length
            )
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 質問
            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=record.question if record.question else "−",
            )
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 回答
            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=record.answer if record.answer else "−",
            )
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 一致FAQ_行
            faq_row_val = (
                "−"
                if record.matched_faq_row is None
                else record.matched_faq_row
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=faq_row_val)
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 一致FAQ_質問
            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=record.matched_faq_question or "−",
            )
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

            # 一致FAQ_回答
            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=record.matched_faq_answer or "−",
            )
            apply_cell_style(cell, fill=fill, is_group_start=is_group_start)
            col_idx += 1

        # 列幅設定（P1.5削除、生_対応結果_文字数追加）
        widths = [
            8,  # 元idx
            10,  # P0_GID
            10,  # P0_類似度
            10,  # P2_類似度
            12,  # P3-1_類似度
            12,  # P3-2_類似度
            10,  # 最終GID
            18,  # 最終結果
            8,  # 信頼度
            40,  # 生_概要
            50,  # 生_対応結果
            12,  # 生_対応結果_文字数
            45,  # 質問
            55,  # 回答
            10,  # 一致FAQ_行
            45,  # 一致FAQ_質問
            55,  # 一致FAQ_回答
        ]
        set_column_widths(ws, widths)

        # 行高さ設定
        ws.row_dimensions[1].height = 25
        for row in range(2, len(records) + 2):
            ws.row_dimensions[row].height = 50

    def save(self) -> str:
        """
        Excelファイルを保存

        Returns:
            保存したファイルのパス
        """
        os.makedirs(self.output_dir, exist_ok=True)
        filepath = os.path.join(self.output_dir, self.get_filename())
        self.wb.save(filepath)
        print(f"✅ 検証Excel出力: {filepath}")
        return filepath


# =============================================================================
# ユーティリティ関数
# =============================================================================
def get_adopted_indices(groups: List[DuplicateGroup]) -> List[int]:
    """
    採用されたインデックスのリストを取得

    Args:
        groups: グループリスト

    Returns:
        採用されたoriginal_idxのリスト
    """
    adopted = []
    for group in groups:
        for candidate in group.candidates:
            if candidate.is_adopted:
                adopted.append(candidate.original_idx)
    return adopted
