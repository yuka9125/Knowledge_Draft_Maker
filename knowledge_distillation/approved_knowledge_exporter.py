#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
レビュー済みExcelから承認済みKnowledgeを出力する。
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


SHEET_NAME = "最終ナレッジ候補一覧"
APPROVED_REVIEW_RESULT = "採用"
JST = timezone(timedelta(hours=9))


def _cell_text(value: Any) -> str:
    """Excelセル値をJSON向け文字列に整える。"""
    if value is None:
        return ""
    return str(value).strip()


def _build_header_index(ws) -> Dict[str, int]:
    """1行目のヘッダー名から列番号へのマップを作る。"""
    headers: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        header = _cell_text(cell.value)
        if header:
            headers[header] = col_idx
    return headers


def _required_value(ws, row_idx: int, headers: Dict[str, int], header: str) -> str:
    """必須列の値を取得する。"""
    if header not in headers:
        raise ValueError(f"必須列が見つかりません: {header}")
    return _cell_text(ws.cell(row=row_idx, column=headers[header]).value)


def _optional_value(ws, row_idx: int, headers: Dict[str, int], header: str) -> str:
    """任意列の値を取得する（列が無ければ空文字）。"""
    if header not in headers:
        return ""
    return _cell_text(ws.cell(row=row_idx, column=headers[header]).value)


EXISTING_FAQ_ID_HEADER = "既存FAQ_ID"


def load_approved_knowledge_from_excel(
    excel_path_or_file: str | Path | Any,
    approved_at: str | None = None,
    use_existing_faq_id: bool = False,
) -> List[Dict[str, str]]:
    """レビュー結果が「採用」の行だけ approved_knowledge 形式で返す。

    use_existing_faq_id=True のとき、既存FAQ_ID が入っている行（＝既存FAQの更新）は
    knowledge_id を既存FAQ_ID に揃える。後段の upsert マージで既存エントリに上書きされる。
    """
    wb = load_workbook(excel_path_or_file, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {SHEET_NAME}")

    ws = wb[SHEET_NAME]
    headers = _build_header_index(ws)
    approved_at_value = approved_at or datetime.now(JST).isoformat(timespec="seconds")

    approved_items: List[Dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        review_result = _required_value(ws, row_idx, headers, "レビュー結果")
        if review_result != APPROVED_REVIEW_RESULT:
            continue

        knowledge_id = _required_value(ws, row_idx, headers, "ナレッジID")
        question = _required_value(ws, row_idx, headers, "候補_質問")
        answer = _required_value(ws, row_idx, headers, "候補_回答")
        category = _required_value(ws, row_idx, headers, "カテゴリ")
        if not knowledge_id or not question or not answer:
            continue

        if use_existing_faq_id:
            existing_faq_id = _optional_value(
                ws, row_idx, headers, EXISTING_FAQ_ID_HEADER
            )
            if existing_faq_id:
                # 既存FAQの更新 → 既存IDに揃えて上書き対象にする
                knowledge_id = existing_faq_id

        approved_items.append(
            {
                "knowledge_id": knowledge_id,
                "question": question,
                "answer": answer,
                "category": category,
                "approved_status": "approved",
                "approved_at": approved_at_value,
            }
        )

    return approved_items


def merge_approved_knowledge(
    base_items: List[Dict[str, str]],
    incoming_items: List[Dict[str, str]],
) -> List[Dict[str, str]]:
    """knowledge_id をキーに upsert マージする。

    - incoming が base と同じ knowledge_id を持てば**上書き**（既存FAQ更新）
    - 新しい knowledge_id は**末尾に追加**（新規）
    - base の並び順を保ち、新規は出現順で後ろに足す
    """
    merged: Dict[str, Dict[str, str]] = {}
    order: List[str] = []
    for item in list(base_items) + list(incoming_items):
        kid = str(item.get("knowledge_id", "")).strip()
        if not kid:
            continue
        if kid not in merged:
            order.append(kid)
        merged[kid] = item  # 後勝ち（incoming が base を上書き）
    return [merged[kid] for kid in order]


def _read_json_list(path: str | Path) -> List[Dict[str, str]]:
    """既存の approved_knowledge.json を読む（無い/壊れていれば空）。"""
    target = Path(path)
    if not target.exists():
        return []
    try:
        with target.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return []
    if not isinstance(data, list):
        return []
    return [item for item in data if isinstance(item, dict)]


def export_approved_knowledge_from_excel(
    excel_path_or_file: str | Path | Any,
    output_path: str | Path,
    approved_at: str | None = None,
    base_path: str | Path | None = None,
    merge: bool = True,
) -> Path:
    """レビュー済みExcelから approved_knowledge.json を出力する。

    merge=True（既定）：既存の approved_knowledge.json（base_path、未指定なら output_path）に
    対して knowledge_id で upsert する。既存FAQの更新は既存IDに上書き、新規は追加するため
    重複が出にくい。
    merge=False：従来どおり「採用行のスナップショット」を上書き出力する。
    """
    approved_items = load_approved_knowledge_from_excel(
        excel_path_or_file=excel_path_or_file,
        approved_at=approved_at,
        use_existing_faq_id=merge,
    )

    if merge:
        base_source = base_path if base_path is not None else output_path
        base_items = _read_json_list(base_source)
        result_items = merge_approved_knowledge(base_items, approved_items)
    else:
        result_items = approved_items

    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8") as f:
        json.dump(result_items, f, ensure_ascii=False, indent=2)
    return target


def main() -> None:
    parser = argparse.ArgumentParser(
        description="レビュー済みExcelから approved_knowledge.json を出力します。"
    )
    parser.add_argument("excel_path", help="FAQ_final_result.xlsx のパス")
    parser.add_argument(
        "-o",
        "--output",
        default="data/approved_knowledge.json",
        help="出力先JSONパス",
    )
    parser.add_argument(
        "--base",
        default=None,
        help="upsertのベースとする既存JSON（未指定なら--outputを使う）",
    )
    parser.add_argument(
        "--no-merge",
        dest="merge",
        action="store_false",
        help="upsertせず採用行のスナップショットで上書きする",
    )
    args = parser.parse_args()

    output_path = export_approved_knowledge_from_excel(
        excel_path_or_file=args.excel_path,
        output_path=args.output,
        base_path=args.base,
        merge=args.merge,
    )
    print(f"approved_knowledge出力: {output_path}")


if __name__ == "__main__":
    main()
